# -*- coding: utf-8 -*-
"""
Migración "Correlativos Perfiles y Moldes" - Paso A (solo lectura).

Lee el Excel maestro (docs/CORRELATIVOS PERFILES _ MOLDES (PLA-DES-CPM-S).xlsx),
limpia placeholders/fórmulas visuales, preserva anomalías históricas sin corregirlas,
y produce:
  - cpm_perfiles.json  : filas listas para insertar en cpm_perfiles
  - cpm_moldes.json    : filas listas para insertar en cpm_moldes (repetitivo + no
                          repetitivo + histórico/obsoleto)
  - reporte_migracion.md : reporte con conteos, anomalías y siguiente correlativo por tipo

No escribe en ninguna base de datos. Reproducible: se puede correr las veces que se
quiera, siempre genera el mismo resultado a partir del mismo Excel.
"""
import openpyxl
import re
import json
import sys
import datetime
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = r"C:\Users\dcarrasco\Desktop\Proyectos\workspacefaret\docs\CORRELATIVOS PERFILES _ MOLDES (PLA-DES-CPM-S).xlsx"
OUT_DIR = r"C:\Users\dcarrasco\Desktop\Proyectos\workspacefaret\tools\migracion-perfiles-moldes"
IMPORT_BATCH = "MIGRACION_2026-08-13"

PLACEHOLDER_RE = re.compile(r'^_+$')
DATE_RE_YY = re.compile(r'^\d{1,2}/\d{1,2}/\d{2}$')
DATE_RE_YYYY = re.compile(r'^\d{1,2}/\d{1,2}/\d{4}$')
ERROR_VALUES = {'#VALUE!', '#REF!', '#N/A', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!'}


def clean_text(v):
    """Placeholder Excel (____) o vacio -> None. Si no, trim."""
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or PLACEHOLDER_RE.match(s):
        return None
    if s in ERROR_VALUES:
        return None
    return s


def parse_fecha(v):
    """Devuelve (fecha_iso_o_None, raw_preservado_o_None, es_anomalia)."""
    if v is None:
        return None, None, False
    if isinstance(v, datetime.datetime):
        if 2010 <= v.year <= 2030:
            return v.date().isoformat(), None, False
        return None, str(v), True
    s = str(v).strip()
    if s == '' or PLACEHOLDER_RE.match(s):
        return None, None, False
    if s in ERROR_VALUES:
        return None, s, True
    if DATE_RE_YY.match(s):
        try:
            return datetime.datetime.strptime(s, '%d/%m/%y').date().isoformat(), None, False
        except ValueError:
            return None, s, True
    if DATE_RE_YYYY.match(s):
        try:
            return datetime.datetime.strptime(s, '%d/%m/%Y').date().isoformat(), None, False
        except ValueError:
            return None, s, True
    # formatos con guion tipo 13-08-24
    if re.match(r'^\d{1,2}-\d{1,2}-\d{2,4}$', s):
        for fmt in ('%d-%m-%y', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(s, fmt).date().isoformat(), None, False
            except ValueError:
                pass
    return None, s, True


def _valor_presente(v):
    if v is None:
        return False
    s = str(v).strip()
    return s != '' and not PLACEHOLDER_RE.match(s) and s not in ERROR_VALUES


def fila_tiene_dato_real(cliente_raw, *otros_valores):
    """CLIENTE es la señal principal de 'fila real' (validado contra los conteos
    reales del Excel: 8.068/1.002/738/171). Pero un puñado de filas de PERFILES
    tienen CLIENTE vacío por un hueco de captura y sin embargo sí tienen
    DESCRIPCIÓN/FECHA reales (ver fila 5029, P005023) - esas también se importan
    para no perder información real, y quedan marcadas en el reporte. Ojo: para
    Moldes NO se usa LISTO/ENTREGA como señal aquí, porque las filas prellenadas
    vacías solo tienen 'PENDIENTE' en esa columna (ver M001003-M001015)."""
    if _valor_presente(cliente_raw):
        return True
    return any(_valor_presente(v) for v in otros_valores)


class Reporte:
    def __init__(self):
        self.secciones = {}

    def seccion(self, nombre):
        self.secciones[nombre] = {}
        return self.secciones[nombre]


def migrar():
    print("Cargando workbook...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print("OK\n")

    reporte = Reporte()
    perfiles_out = []
    moldes_out = []

    # ---------------- PERFILES ----------------
    ws = wb['PERFILES']
    header_row = 4
    total_filas = ws.max_row - header_row
    real_count = 0
    placeholder_count = 0
    duplicados = Counter()
    codigos_nulo = []
    codigos_version = []
    caja_no_numerico = 0
    unid_no_numerico = 0
    fechas_anomalas = []
    estado_ref_error = []
    ultimo_p_num = 0
    cliente_vacio_importado = []

    for r in range(header_row + 1, ws.max_row + 1):
        codigo = clean_text(ws.cell(row=r, column=2).value)
        cliente_raw = ws.cell(row=r, column=3).value
        desc_raw = ws.cell(row=r, column=5).value
        fecha_raw_val = ws.cell(row=r, column=7).value
        if not fila_tiene_dato_real(cliente_raw, desc_raw, fecha_raw_val):
            placeholder_count += 1
            continue
        real_count += 1
        if not _valor_presente(cliente_raw):
            cliente_vacio_importado.append((r, codigo, clean_text(desc_raw)))

        if codigo:
            duplicados[codigo] += 1
            if 'NULO' in codigo.upper():
                codigos_nulo.append((r, codigo))
            if '-V' in codigo.upper():
                codigos_version.append((r, codigo))
            m = re.match(r'^P0*(\d+)$', codigo)
            if m:
                ultimo_p_num = max(ultimo_p_num, int(m.group(1)))

        fecha_iso, fecha_raw, es_anom = parse_fecha(ws.cell(row=r, column=7).value)
        if es_anom:
            fechas_anomalas.append((r, fecha_raw))

        numero_caja = clean_text(ws.cell(row=r, column=10).value)
        if numero_caja is not None:
            try:
                float(numero_caja)
            except ValueError:
                caja_no_numerico += 1

        unidades_por_caja = clean_text(ws.cell(row=r, column=11).value)
        if unidades_por_caja is not None:
            try:
                float(unidades_por_caja)
            except ValueError:
                unid_no_numerico += 1

        estado_val = ws.cell(row=r, column=13).value
        estado = clean_text(estado_val)
        if estado_val == '#REF!':
            estado_ref_error.append(r)
            estado = None

        perfiles_out.append({
            'numero_perfil': codigo,
            'cliente': clean_text(cliente_raw),
            'medidas': clean_text(ws.cell(row=r, column=4).value),
            'descripcion': clean_text(ws.cell(row=r, column=5).value),
            'numero_desarrollo': clean_text(ws.cell(row=r, column=6).value),
            'fecha': fecha_iso,
            'fecha_raw': fecha_raw,
            'rubro': clean_text(ws.cell(row=r, column=8).value),
            'operador': clean_text(ws.cell(row=r, column=9).value),
            'numero_caja': numero_caja,
            'unidades_por_caja': unidades_por_caja,
            'status': clean_text(ws.cell(row=r, column=12).value),
            'estado': estado,
            'perfil_nuevo': clean_text(ws.cell(row=r, column=14).value),
            'source_sheet': 'PERFILES',
            'source_row': r,
            'import_batch': IMPORT_BATCH,
        })

    dupes = {k: v for k, v in duplicados.items() if v > 1}
    sec = reporte.seccion('PERFILES')
    sec['filas_escaneadas'] = total_filas
    sec['filas_importadas'] = real_count
    sec['filas_ignoradas_placeholder'] = placeholder_count
    sec['duplicados_historicos'] = dupes
    sec['codigos_nulo'] = codigos_nulo
    sec['codigos_con_version'] = codigos_version
    sec['caja_valores_no_numericos'] = caja_no_numerico
    sec['unidades_valores_no_numericos'] = unid_no_numerico
    sec['fechas_no_interpretables'] = fechas_anomalas
    sec['estado_con_error_formula'] = estado_ref_error
    sec['cliente_vacio_pero_importado'] = cliente_vacio_importado
    sec['ultimo_codigo_real'] = f'P{ultimo_p_num:06d}'
    sec['siguiente_correlativo'] = f'P{ultimo_p_num + 1:06d}'

    print(f"PERFILES: {real_count} reales / {total_filas} escaneadas, "
          f"siguiente correlativo P{ultimo_p_num + 1:06d}")

    # ---------------- MOLDES NUEVOS (REPETITIVO) ----------------
    def migrar_moldes(sheet_name, tipo_molde, es_obsoleto, tiene_entrega, prefijo_regex):
        ws = wb[sheet_name]
        real_count = 0
        placeholder_count = 0
        duplicados = Counter()
        operador_como_fecha = []
        fechas_ingreso_anomalas = []
        fechas_entrega_anomalas = []
        codigos_especiales = []
        ultimo_num = 0

        col = {'cod': 1, 'cliente': 2, 'rubro': 3, 'producto': 4}
        if tiene_entrega:
            # MOLDE NO REPETITVO / OBSOLETO: perfil(5) ingreso(6) listo(7) entrega(8) operador(9) np(10) comentarios(11)
            col.update(perfil=5, ingreso=6, listo=7, entrega=8, operador=9, np=10, comentarios=11)
        else:
            # MOLDES NUEVOS: np(5) perfil(6) ingreso(7) listo(8) operador(9) comentarios(10)
            col.update(np=5, perfil=6, ingreso=7, listo=8, operador=9, comentarios=None)
            col['comentarios'] = 10

        for r in range(5, ws.max_row + 1):
            codigo = clean_text(ws.cell(row=r, column=col['cod']).value)
            cliente_raw = ws.cell(row=r, column=col['cliente']).value
            producto_raw = ws.cell(row=r, column=col['producto']).value
            perfil_raw = ws.cell(row=r, column=col['perfil']).value
            ingreso_raw = ws.cell(row=r, column=col['ingreso']).value
            # Ojo: 'listo'/'entrega' deliberadamente NO se usan como señal - las filas
            # prellenadas vacías solo tienen 'PENDIENTE' ahi (ver M001003-M001015).
            if not fila_tiene_dato_real(cliente_raw, producto_raw, perfil_raw, ingreso_raw):
                placeholder_count += 1
                continue
            real_count += 1

            if codigo:
                duplicados[codigo] += 1
                m = re.match(prefijo_regex, codigo)
                if m:
                    ultimo_num = max(ultimo_num, int(m.group(1)))
                else:
                    codigos_especiales.append((r, codigo))

            fecha_ingreso_iso, fecha_ingreso_raw, anom_i = parse_fecha(ws.cell(row=r, column=col['ingreso']).value)
            if anom_i:
                fechas_ingreso_anomalas.append((r, fecha_ingreso_raw))

            fecha_entrega_iso = fecha_entrega_raw = None
            if tiene_entrega:
                fecha_entrega_iso, fecha_entrega_raw, anom_e = parse_fecha(ws.cell(row=r, column=col['entrega']).value)
                if anom_e:
                    fechas_entrega_anomalas.append((r, fecha_entrega_raw))

            operador_val = ws.cell(row=r, column=col['operador']).value
            if isinstance(operador_val, datetime.datetime):
                operador_como_fecha.append((r, codigo, str(operador_val)))
                operador = str(operador_val)  # se preserva tal cual, sin inventar el nombre
            else:
                operador = clean_text(operador_val)

            np_val = ws.cell(row=r, column=col['np']).value
            # NP (1RA ENTRADA): se preserva tal cual (texto/fecha/numero mezclados en el Excel
            # original) sin reinterpretar - ver CLAUDE.md seccion 9 del requerimiento.
            if np_val is None:
                np_texto = None
            elif isinstance(np_val, datetime.datetime):
                np_texto = np_val.date().isoformat()
            else:
                np_texto = clean_text(np_val)

            moldes_out.append({
                'tipo_molde': tipo_molde,
                'codigo': codigo,
                'cliente': clean_text(cliente_raw),
                'rubro': clean_text(ws.cell(row=r, column=col['rubro']).value),
                'producto': clean_text(ws.cell(row=r, column=col['producto']).value),
                'np_primera_entrada': np_texto,
                'perfil': clean_text(ws.cell(row=r, column=col['perfil']).value),
                'fecha_ingreso': fecha_ingreso_iso,
                'fecha_ingreso_raw': fecha_ingreso_raw,
                'listo_estado': clean_text(ws.cell(row=r, column=col['listo']).value),
                'fecha_entrega': fecha_entrega_iso,
                'fecha_entrega_raw': fecha_entrega_raw,
                'operador': operador,
                'comentarios': clean_text(ws.cell(row=r, column=col['comentarios']).value),
                'es_obsoleto': 1 if es_obsoleto else 0,
                'source_sheet': sheet_name,
                'source_row': r,
                'import_batch': IMPORT_BATCH,
            })

        dupes = {k: v for k, v in duplicados.items() if v > 1}
        return {
            'filas_escaneadas': ws.max_row - 4,
            'filas_importadas': real_count,
            'filas_ignoradas_placeholder': placeholder_count,
            'duplicados': dupes,
            'codigos_especiales_formato': codigos_especiales,
            'operador_reemplazado_por_fecha': operador_como_fecha,
            'fechas_ingreso_no_interpretables': fechas_ingreso_anomalas,
            'fechas_entrega_no_interpretables': fechas_entrega_anomalas,
            'ultimo_numero_real': ultimo_num,
        }

    rep_nuevos = migrar_moldes('MOLDES NUEVOS(PLA-DES-CPM)', 'REPETITIVO', False, False, r'^M0*(\d+)$')
    reporte.secciones['MOLDES_REPETITIVO'] = rep_nuevos
    reporte.secciones['MOLDES_REPETITIVO']['siguiente_correlativo'] = f"M{rep_nuevos['ultimo_numero_real']+1:06d}"
    print(f"MOLDES REPETITIVO: {rep_nuevos['filas_importadas']} reales, "
          f"siguiente M{rep_nuevos['ultimo_numero_real']+1:06d}")

    rep_nr = migrar_moldes('MOLDE NO REPETITVO(PLA-DES-CPM)', 'NO_REPETITIVO', False, True, r'^NR0*(\d+)$')
    reporte.secciones['MOLDES_NO_REPETITIVO'] = rep_nr
    reporte.secciones['MOLDES_NO_REPETITIVO']['siguiente_correlativo'] = f"NR{rep_nr['ultimo_numero_real']+1:06d}"
    print(f"MOLDES NO REPETITIVO: {rep_nr['filas_importadas']} reales, "
          f"siguiente NR{rep_nr['ultimo_numero_real']+1:06d}")

    rep_obs = migrar_moldes('MOLDES (OBSOLETO)', 'HISTORICO', True, True, r'^MT0*(\d+)$')
    reporte.secciones['MOLDES_OBSOLETO'] = rep_obs
    print(f"MOLDES OBSOLETO: {rep_obs['filas_importadas']} reales (formatos de codigo mixtos, "
          f"no genera correlativo)")

    # ---------------- Hojas NO migradas (solo inventario) ----------------
    hojas_2023 = [n for n in wb.sheetnames if re.match(r'^\d{2}\.-', n)]
    inventario = {}
    for hn in hojas_2023:
        wsx = wb[hn]
        inventario[hn] = wsx.max_row
    reporte.secciones['HOJAS_NO_MIGRADAS'] = {
        'solicitud_desarrollos_2023': inventario,
        'otras_hojas_auxiliares_no_migradas': ['Programa', 'PLANTILLA'],
        'motivo': 'Proceso distinto (Solicitud de Desarrollos 2023), no correlativos Perfiles/Moldes',
    }

    # ---------------- Escribir salidas ----------------
    with open(f"{OUT_DIR}\\cpm_perfiles.json", 'w', encoding='utf-8') as f:
        json.dump(perfiles_out, f, ensure_ascii=False, indent=None)
    with open(f"{OUT_DIR}\\cpm_moldes.json", 'w', encoding='utf-8') as f:
        json.dump(moldes_out, f, ensure_ascii=False, indent=None)

    escribir_reporte(reporte, f"{OUT_DIR}\\reporte_migracion.md")

    print(f"\nOK: {len(perfiles_out)} perfiles + {len(moldes_out)} moldes escritos en JSON.")
    print(f"Reporte: {OUT_DIR}\\reporte_migracion.md")


def escribir_reporte(reporte, path):
    lines = []
    lines.append("# Reporte de migración — Correlativos Perfiles y Moldes")
    lines.append(f"\nGenerado a partir de: `{XLSX_PATH}`")
    lines.append(f"\nLote de importación: `{IMPORT_BATCH}`\n")

    p = reporte.secciones['PERFILES']
    lines.append("## PERFILES")
    lines.append(f"- Filas escaneadas: {p['filas_escaneadas']}")
    lines.append(f"- Filas importadas (con CLIENTE real): **{p['filas_importadas']}**")
    lines.append(f"- Filas ignoradas por ser placeholder/vacías: {p['filas_ignoradas_placeholder']}")
    lines.append(f"- Último código real: {p['ultimo_codigo_real']} → siguiente correlativo: **{p['siguiente_correlativo']}**")
    lines.append(f"- Códigos duplicados históricos ({len(p['duplicados_historicos'])}): {p['duplicados_historicos']}")
    lines.append(f"- Código NULO: {p['codigos_nulo']}")
    lines.append(f"- Códigos con versión (-V2/-V3): {len(p['codigos_con_version'])} filas")
    lines.append(f"- N° DE CAJA con valores no numéricos: {p['caja_valores_no_numericos']} filas")
    lines.append(f"- UNID POR CAJA con valores no numéricos: {p['unidades_valores_no_numericos']} filas")
    lines.append(f"- Fechas no interpretables (quedan NULL, valor original preservado en fecha_raw): {len(p['fechas_no_interpretables'])} filas")
    for r, raw in p['fechas_no_interpretables'][:50]:
        lines.append(f"    - fila {r}: {raw!r}")
    lines.append(f"- ESTADO con error de fórmula (#REF!, guardado como NULL): filas {p['estado_con_error_formula']}")
    lines.append(f"- Filas con CLIENTE vacío pero importadas por tener otro dato real (desc/fecha): {len(p['cliente_vacio_pero_importado'])}")
    for r, cod, desc in p['cliente_vacio_pero_importado']:
        lines.append(f"    - fila {r} ({cod}): descripcion={desc!r}")

    for key, titulo in [('MOLDES_REPETITIVO', 'MOLDES NUEVOS (Repetitivo)'),
                         ('MOLDES_NO_REPETITIVO', 'MOLDE NO REPETITIVO'),
                         ('MOLDES_OBSOLETO', 'MOLDES (OBSOLETO, histórico)')]:
        m = reporte.secciones[key]
        lines.append(f"\n## {titulo}")
        lines.append(f"- Filas escaneadas: {m['filas_escaneadas']}")
        lines.append(f"- Filas importadas (con CLIENTE real): **{m['filas_importadas']}**")
        lines.append(f"- Filas ignoradas por ser placeholder/vacías: {m['filas_ignoradas_placeholder']}")
        if 'siguiente_correlativo' in m:
            lines.append(f"- Último número real: {m['ultimo_numero_real']} → siguiente correlativo: **{m['siguiente_correlativo']}**")
        lines.append(f"- Duplicados: {m['duplicados']}")
        lines.append(f"- Códigos con formato especial/histórico (no matchean el patrón moderno): {len(m['codigos_especiales_formato'])} filas")
        for r, c in m['codigos_especiales_formato'][:30]:
            lines.append(f"    - fila {r}: {c!r}")
        if m['operador_reemplazado_por_fecha']:
            lines.append(f"- ANOMALÍA: OPERADOR contiene una fecha en vez de un nombre ({len(m['operador_reemplazado_por_fecha'])} filas) — preservado tal cual, no se inventó el operador:")
            for r, cod, val in m['operador_reemplazado_por_fecha']:
                lines.append(f"    - fila {r} ({cod}): {val}")
        lines.append(f"- Fechas de ingreso no interpretables: {len(m['fechas_ingreso_no_interpretables'])} filas")
        for r, raw in m['fechas_ingreso_no_interpretables'][:30]:
            lines.append(f"    - fila {r}: {raw!r}")
        if m['fechas_entrega_no_interpretables']:
            lines.append(f"- Fechas de entrega no interpretables: {len(m['fechas_entrega_no_interpretables'])} filas")
            for r, raw in m['fechas_entrega_no_interpretables'][:30]:
                lines.append(f"    - fila {r}: {raw!r}")

    h = reporte.secciones['HOJAS_NO_MIGRADAS']
    lines.append("\n## Hojas NO migradas (solo inventariadas)")
    lines.append("Corresponden al proceso 'Solicitud de Desarrollos 2023', distinto de Correlativos Perfiles/Moldes:")
    for hoja, filas in h['solicitud_desarrollos_2023'].items():
        lines.append(f"- {hoja}: {filas} filas (máx. fila con datos, sin depurar placeholders)")
    lines.append(f"- Otras hojas auxiliares no migradas: {h['otras_hojas_auxiliares_no_migradas']} (basadas en fórmulas / plantilla, no son datos operacionales)")

    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


if __name__ == '__main__':
    migrar()
